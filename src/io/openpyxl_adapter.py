from openpyxl.styles import NamedStyle, Side, Border, Font, Alignment, PatternFill

from src.formats.styles.base import TableStyle


class OpenPyXLStyleAdapter:
    @staticmethod
    def convert(table_style: TableStyle) -> NamedStyle:
        style = NamedStyle(name=table_style.name)
        style.font = OpenPyXLStyleAdapter._convert_font(table_style)
        style.border = OpenPyXLStyleAdapter._convert_border(table_style)
        style.alignment = OpenPyXLStyleAdapter._convert_alignment(table_style)
        style.fill = OpenPyXLStyleAdapter._convert_fill(table_style)
        return style

    @staticmethod
    def _convert_font(table_style: TableStyle) -> Font:
        font_weights = table_style.font.weight or []
        weight_lower = [w.lower() for w in font_weights]

        return Font(
            name=table_style.font.name,
            size=table_style.font.size,
            bold="bold" in weight_lower,
            italic="italic" in weight_lower,
            underline="single" if "underline" in weight_lower else None,
            color=f"00{table_style.font.color.value}",
        )

    @staticmethod
    def _convert_border(table_style: TableStyle) -> Border:
        if not table_style.border:
            return Border()

        thickness_map = {1: "thin", 2: "medium", 3: "thick"}
        border_style = thickness_map.get(table_style.border.thickness, "thin")
        side = Side(style=border_style, color=f"00{table_style.border.color.value}")

        sides = [s.value for s in table_style.border.sides]
        return Border(
            left=side if "left" in sides else None,
            right=side if "right" in sides else None,
            top=side if "top" in sides else None,
            bottom=side if "bottom" in sides else None,
        )

    @staticmethod
    def _convert_alignment(table_style: TableStyle) -> Alignment:
        return Alignment(horizontal=table_style.alignment.value, vertical="center")

    @staticmethod
    def _convert_fill(table_style: TableStyle) -> PatternFill:
        if not table_style.fill_color:
            return PatternFill()

        return PatternFill(
            fill_type="solid", start_color=f"00{table_style.fill_color.value}"
        )
