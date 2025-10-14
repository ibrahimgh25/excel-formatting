import pandas as pd
from openpyxl import load_workbook


def read_excel_sheet(file_path: str, sheet_name: str | None = None) -> pd.DataFrame:
    try:
        if sheet_name is None:
            df = pd.read_excel(file_path)
        else:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        return df
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    except ValueError as e:
        raise ValueError(f"Error reading sheet '{sheet_name}': {str(e)}")


def get_sheet_names(file_path: str) -> list[str]:
    try:
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found: {file_path}")
