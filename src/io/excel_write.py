from typing import Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from src.formats.styles.base import TableStyle
from src.formats.styles.builtin import get_style
from .openpyxl_adapter import OpenPyXLStyleAdapter


def _add_styles_to_workbook(wb: Workbook, style_names: list[str]) -> Workbook:
    for style_name in style_names:
        try:
            table_style = get_style(style_name)
            if isinstance(table_style, TableStyle):
                named_style = OpenPyXLStyleAdapter.convert(table_style)
                wb.add_named_style(named_style)
        except ValueError:
            pass
    return wb


def _adjust_column_width(ws: Worksheet, col_index: int, width: float) -> None:
    assert width >= 0, "A negative width isn't allowed"
    letter = get_column_letter(col_index + 1)
    if width == 0:
        ws.column_dimensions[letter].auto_size = True
    else:
        ws.column_dimensions[letter].width = width


def save_dataframe_to_excel(
    df: pd.DataFrame,
    file_path: str,
    sheet_name: str,
    column_styles: Optional[list[dict[str, str]]] = None,
    column_widths: Optional[list[float]] = None,
    hide_gridlines: bool = True,
) -> None:
    try:
        wb = load_workbook(file_path)
        existing_sheets = wb.sheetnames
        index = None

        if sheet_name in existing_sheets:
            index = existing_sheets.index(sheet_name)
            sheet_to_remove = wb[sheet_name]
            wb.remove(sheet_to_remove)

        ws = wb.create_sheet(sheet_name, index)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    if hide_gridlines:
        ws.sheet_view.showGridLines = False

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    if column_styles:
        assert len(column_styles) == len(
            df.columns
        ), "The number of styles and number of columns in the dataframe must match"

        style_names = set()
        for col_style in column_styles:
            style_names.add(col_style["Header"])
            style_names.add(col_style["Body"])

        wb = _add_styles_to_workbook(wb, list(style_names))

        for col_nb, col in enumerate(ws.iter_cols()):
            for row_nb, cell in enumerate(col):
                if row_nb == 0:
                    cell.style = column_styles[col_nb]["Header"]
                else:
                    cell.style = column_styles[col_nb]["Body"]

    if column_widths:
        assert len(column_widths) == len(
            df.columns
        ), "The number of column widths and number of columns in the dataframe must match"

        for index, width in enumerate(column_widths):
            _adjust_column_width(ws, index, width)

    try:
        wb.save(file_path)
    except PermissionError:
        print(
            f"{file_path} may be still open, please close it and press Enter to try again"
        )
        input()
        wb.save(file_path)


def create_style_dict(header_style: str, body_style: str) -> dict[str, str]:
    assert isinstance(header_style, str), "header_style should be a string"
    assert isinstance(body_style, str), "body_style should be a string"
    return {"Header": header_style, "Body": body_style}
